"""
Coffee Aggregator — Scraper Pipeline Entry Point

Usage (from the Scraper/ directory):
    pip install -r requirements.txt
    python scraper/main.py          # CLI mode — prints to console, writes files

    # OR via FastAPI (see server/app.py):
    uvicorn server.app:app --port 8000

Outputs written to output/:
    products.json          — normalized product dataset
    products.xlsx          — same data formatted for manual review
    scrape_log.json        — per-roaster scrape results and errors
    images_manifest.json   — flat list of all product image URLs
"""

# ── Happy-eyeballs DNS patch ─────────────────────────────────────────
#
# This file runs as a subprocess spawned by
# `services/scrape_runner.py` in the catalog-ops backend. The API
# server installs `services.http_client.install_urllib3_patch()` at
# boot to fix the multi-A-record / poisoned-resolver blackhole
# problem (see services/http_client.py). The subprocess gets a fresh
# Python interpreter and DOES NOT inherit that patch — so on
# networks where the local resolver returns blackhole IPs (Reserved,
# Mindful, anything CNAME'd to shops.myshopify.com on some ISPs),
# the scraper's first `requests.get` against the storefront times
# out and the scrape returns 0 products silently. Mindful Coffee
# Roaster's "scraped=0" failure today traced back to this exact
# subprocess-vs-server-process gap.
#
# Install the same patch at the top of the subprocess so every
# `requests` / `urllib3` call inherits happy-eyeballs + DoH
# fallback. Idempotent — safe to call from main() too.
import sys as _sys
import os as _os
_backend = _os.path.abspath(_os.path.join(
    _os.path.dirname(__file__), "..", "..", "Community",
    "coffee-community-api",
))
if _backend not in _sys.path:
    _sys.path.insert(0, _backend)
try:
    from services.http_client import install_urllib3_patch  # type: ignore
    install_urllib3_patch()
except Exception as _e:
    # Don't block the scraper if the patch can't load — the legacy
    # CLI-only mode (no catalog-ops backend in sight) must still run.
    _sys.stderr.write(
        f"[scraper] WARN: happy-eyeballs patch unavailable: {_e!s}\n"
    )

import datetime
import json
import os
import sys
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

# Ensure scraper/ directory is on the path so sibling modules resolve
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from platform_detector import confirm_platform
from shopify_scraper import scrape_shopify
from woocommerce_scraper import scrape_woocommerce
from custom_scraper import scrape_custom
from filters import is_coffee_product, is_confirmed_coffee_bean
from normalizer import (
    normalize_shopify_product,
    normalize_woocommerce_product,
    normalize_custom_product,
)
from utils import variants_to_display

# ── Paths ─────────────────────────────────────────────────────────────────────

_BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# Per-roaster parallelism: when SCRAPER_INPUT_PATH / SCRAPER_OUTPUT_DIR
# are set (by the FastAPI orchestrator's per-slug workspace), use them.
# Otherwise fall back to the legacy global paths so standalone CLI runs
# still work (`python scraper/main.py` from the Scraper/ dir).
_INPUT_PATH = os.environ.get(
    "SCRAPER_INPUT_PATH",
    os.path.join(_BASE_DIR, "input", "verified_roasters_catalog.json"),
)
_OUTPUT_DIR = os.environ.get(
    "SCRAPER_OUTPUT_DIR",
    os.path.join(_BASE_DIR, "output"),
)

# ── Excel formatting constants ────────────────────────────────────────────────

_FILL_YELLOW = PatternFill("solid", fgColor="FFFF00")
_FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
_FILL_ORANGE = PatternFill("solid", fgColor="FFEB9C")
_FILL_RED = PatternFill("solid", fgColor="FFC7CE")
_FILL_HEADER = PatternFill("solid", fgColor="2F5496")
_FONT_HEADER = Font(bold=True, color="FFFFFF")

_COLUMNS = [
    "product_id", "roaster_name", "roaster_slug", "roaster_city",
    "roaster_state", "roaster_lat", "roaster_lng", "roaster_website",
    "coffee_name", "coffee_slug", "roast_level", "tasting_notes",
    "origin", "altitude_masl", "process", "varietal",
    "weight_grams", "price_inr", "price_per_gram", "currency",
    "grind_options", "image_url", "product_url", "available", "variants",
    "tags", "description_raw", "scrape_confidence", "scrape_flags",
    "scraped_at",
]


# ── Quality gate ──────────────────────────────────────────────────────────────

def passes_quality_gate(product: dict) -> bool:
    """
    Strict quality gate — only products meeting ALL criteria pass:
    - High confidence scrape
    - Currently in stock
    - Known roast level (not Unknown/null)
    - Has a valid price and weight
    """
    if product.get("scrape_confidence") != "high":
        return False
    if not product.get("available"):
        return False
    if product.get("roast_level") in ("Unknown", None, ""):
        return False
    if product.get("price_inr") is None:
        return False
    if product.get("weight_grams") is None:
        return False
    return True


# ── Helpers ───────────────────────────────────────────────────────────────────

def _load_catalog(path=None) -> list:
    with open(path or _INPUT_PATH, encoding="utf-8") as f:
        return json.load(f)


def _write_json_atomic(path: str, data) -> None:
    """Write JSON with atomic rename to prevent partial reads."""
    dir_name = os.path.dirname(path)
    fd, tmp_path = tempfile.mkstemp(dir=dir_name, suffix=".tmp")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, path)
    except Exception:
        os.unlink(tmp_path)
        raise


def _cell_value(product: dict, col: str):
    val = product.get(col)
    if col == "variants":
        return variants_to_display(val or [])
    if col in ("tags", "grind_options", "scrape_flags"):
        return ", ".join(str(x) for x in val) if isinstance(val, list) else (val or "")
    if isinstance(val, bool):
        return "Yes" if val else "No"
    if val is None:
        return ""
    return val


def _write_excel(path: str, products: list) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    for col_idx, col_name in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNS))}1"

    confidence_col = _COLUMNS.index("scrape_confidence") + 1
    flags_col = _COLUMNS.index("scrape_flags") + 1

    for row_idx, product in enumerate(products, start=2):
        for col_idx, col_name in enumerate(_COLUMNS, start=1):
            val = _cell_value(product, col_name)
            if col_name == "product_url" and val:
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")
            else:
                ws.cell(row=row_idx, column=col_idx, value=val)

        conf = product.get("scrape_confidence", "")
        conf_cell = ws.cell(row=row_idx, column=confidence_col)
        if conf == "high":
            conf_cell.fill = _FILL_GREEN
        elif conf == "medium":
            conf_cell.fill = _FILL_ORANGE
        elif conf == "low":
            conf_cell.fill = _FILL_RED

        flags = product.get("scrape_flags") or []
        if flags:
            ws.cell(row=row_idx, column=flags_col).fill = _FILL_YELLOW

    for col_idx, col_name in enumerate(_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(col_name),
            max((len(str(_cell_value(p, col_name))) for p in products), default=0),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(path)


def _write_images_manifest(path: str, products: list) -> None:
    urls = sorted({p["image_url"] for p in products if p.get("image_url")})
    _write_json_atomic(path, {"total_images": len(urls), "urls": urls})


def _dedup_products(products: list) -> list:
    seen = set()
    out = []
    for p in products:
        pid = p.get("product_id", "")
        if pid not in seen:
            seen.add(pid)
            out.append(p)
    return out


def _write_outputs(all_products: list, scrape_log: list) -> None:
    """Write all output files (called at end of scrape)."""
    os.makedirs(_OUTPUT_DIR, exist_ok=True)

    _write_json_atomic(os.path.join(_OUTPUT_DIR, "products.json"), all_products)
    print("  products.json ✓")

    _write_excel(os.path.join(_OUTPUT_DIR, "products.xlsx"), all_products)
    print("  products.xlsx ✓")

    _write_json_atomic(os.path.join(_OUTPUT_DIR, "scrape_log.json"), scrape_log)
    print("  scrape_log.json ✓")

    _write_images_manifest(
        os.path.join(_OUTPUT_DIR, "images_manifest.json"), all_products
    )
    print("  images_manifest.json ✓")


# ── Discovery — generic-first (new) + legacy fallback ──────────────────────

def _scrape_legacy_dispatch(roaster: dict, detected: str) -> list:
    """The old per-platform discovery dispatch.

    Kept as an escape hatch (CREMA_LEGACY_DISCOVERY=1) during the
    refactor shakeout — once the generic path proves it's strictly
    >= this one on every roaster in the catalog, this gets deleted.
    """
    if detected == "shopify":
        return scrape_shopify(roaster)
    if detected == "woocommerce":
        woo_products, needs_fallback = scrape_woocommerce(roaster)
        if needs_fallback:
            raw = scrape_custom(roaster)
            for p in raw:
                p["_platform"] = "custom"
            return raw
        return woo_products
    return scrape_custom(roaster)


def _scrape_via_generic_discovery(roaster: dict, detected: str, log_entry: dict) -> list:
    """Generic-first discovery — sitemap + platform augmenter.

    Calls `services.product_discovery.discover` (in the catalog-ops
    backend, on sys.path via the http_client injection at top of
    this file). The sitemap walker handles platform-agnostic URL
    discovery; the per-platform augmenter (when available) attaches
    variant/SKU/price data the sitemap doesn't carry.

    For URLs the augmenter doesn't cover (typical for Wix or for
    Shopify roasters whose shop_url is narrower than the full
    catalog), we route through the existing custom_scraper per-page
    extractor `_scrape_product_page` to fill in title/price/body.
    That path already does the Tier 2-3 JSON-LD + body extraction
    that the rest of the pipeline expects.
    """
    try:
        from services.product_discovery import discover  # type: ignore
    except ImportError as e:
        # Backend module not in scope — fall back to legacy to keep
        # the standalone CLI mode runnable.
        sys.stderr.write(
            f"[scraper] generic discovery unavailable ({e!s}), "
            f"falling back to legacy dispatch\n"
        )
        return _scrape_legacy_dispatch(roaster, detected)

    from custom_scraper import _scrape_product_page  # for sitemap-only enrichment
    from urllib.parse import urlparse as _urlparse

    domain = _urlparse(roaster["website"]).netloc.replace("www.", "")
    result = discover(roaster, log=lambda s: print(f"  [discovery] {s}"))

    log_entry["discovery_urls"] = len(result.urls)
    log_entry["discovery_source_breakdown"] = dict(result.source_breakdown)
    if result.filter_collapsed:
        log_entry["discovery_filter_collapsed"] = True

    raw_products: list = []
    sitemap_only_count = 0

    for dp in result.urls:
        # Augmenter-provided data → use directly. Preserves the
        # platform-shape the existing normalizer expects.
        if dp.augmented:
            shopify_raw = dp.augmented.get("shopify_raw")
            wc_raw = dp.augmented.get("woocommerce_raw")
            if shopify_raw:
                rp = dict(shopify_raw)
                rp["_roaster"] = roaster
                rp["_domain"] = dp.augmented.get("_domain") or domain
                rp["_platform"] = "shopify"
                raw_products.append(rp)
                continue
            if wc_raw:
                rp = dict(wc_raw)
                rp["_roaster"] = roaster
                rp["_domain"] = dp.augmented.get("_domain") or domain
                rp["_platform"] = "woocommerce"
                raw_products.append(rp)
                continue

        # Sitemap-only — fetch the product page and extract via the
        # existing custom-scraper extraction (JSON-LD Product + body
        # text). Same path Wix-platform roasters take today.
        product = _scrape_product_page(dp.url, roaster, domain)
        if product:
            product["_platform"] = "custom"
            raw_products.append(product)
            sitemap_only_count += 1

    log_entry["discovery_sitemap_only_fetched"] = sitemap_only_count
    return raw_products


# ── Per-roaster scrape logic ──────────────────────────────────────────────────

def _scrape_single_roaster(roaster: dict) -> tuple:
    """
    Scrape, filter, and normalize products for a single roaster.
    Returns: (normalized_products: list, log_entry: dict)
    Raises on fatal errors (cloudflare, timeout, etc.)
    """
    name = roaster["name"]
    catalog_platform = roaster.get("platform", "custom").lower()
    log_entry = {
        "roaster": name,
        "website": roaster["website"],
        "platform_catalog": catalog_platform,
        "platform_detected": None,
        "status": "pending",
        "products_found": 0,
        "coffee_products": 0,
        "non_coffee_excluded": 0,
        "confidence_breakdown": {"high": 0, "medium": 0, "low": 0},
        "flags_summary": [],
        "scrape_duration_seconds": 0.0,
        "scraped_at": None,
    }

    t_start = time.time()

    # 1. Detect platform
    detected = confirm_platform(roaster["website"])
    log_entry["platform_detected"] = detected

    if detected != catalog_platform:
        log_entry["platform_mismatch"] = True

    # 2. Scrape — generic-first discovery (sitemap + platform
    # augmenter) by default. Set CREMA_LEGACY_DISCOVERY=1 to fall
    # back to the old per-platform dispatch (kept as escape hatch
    # during the refactor shakeout).
    use_legacy = os.environ.get("CREMA_LEGACY_DISCOVERY") == "1"
    if use_legacy:
        raw_products = _scrape_legacy_dispatch(roaster, detected)
        log_entry["discovery_mode"] = "legacy"
    else:
        raw_products = _scrape_via_generic_discovery(roaster, detected, log_entry)
        log_entry["discovery_mode"] = "generic"

    log_entry["products_found"] = len(raw_products)

    # 3. Filter to coffee only
    coffee_raws = []
    excluded = 0
    for rp in raw_products:
        title = rp.get("title") or rp.get("name") or ""
        ptype = rp.get("product_type") or rp.get("type") or ""
        tags = rp.get("tags") or []
        body = (
            rp.get("body_html")
            or rp.get("short_description")
            or rp.get("description")
            or ""
        )
        is_coffee, is_uncertain = is_coffee_product(
            title=title, product_type=ptype, tags=tags, body_html=body,
        )
        if not is_coffee:
            excluded += 1
            continue
        if is_uncertain:
            rp.setdefault("_extra_flags", []).append("uncertain_category")
        coffee_raws.append(rp)

    log_entry["coffee_products"] = len(coffee_raws)
    log_entry["non_coffee_excluded"] = excluded

    # 4. Normalize
    platform = detected
    normalized = []
    for rp in coffee_raws:
        try:
            if platform == "shopify":
                product = normalize_shopify_product(rp, roaster)
            elif platform == "woocommerce" and rp.get("_platform") != "custom":
                product = normalize_woocommerce_product(rp, roaster)
            else:
                product = normalize_custom_product(rp, roaster)

            if product is None:
                continue

            # Stage 2: structural check — flag products that look uncertain
            # but do NOT drop them; the LLM enrichment pass makes the final call.
            if not is_confirmed_coffee_bean(product):
                product.setdefault("scrape_flags", []).append("needs_llm_review")

            extra = rp.get("_extra_flags", [])
            if extra:
                product["scrape_flags"] = product.get("scrape_flags", []) + extra

            normalized.append(product)
        except Exception as e:
            log_entry.setdefault("product_errors", []).append(str(e))

    # 5. Log metadata
    conf = {"high": 0, "medium": 0, "low": 0}
    flag_counter = {}
    for p in normalized:
        c = p.get("scrape_confidence", "low")
        if c in conf:
            conf[c] += 1
        for f in p.get("scrape_flags", []):
            flag_counter[f] = flag_counter.get(f, 0) + 1

    log_entry["confidence_breakdown"] = conf
    log_entry["flags_summary"] = [
        f"{k}: {v}" for k, v in sorted(flag_counter.items(), key=lambda x: -x[1])
    ]
    log_entry["status"] = "success"
    log_entry["scrape_duration_seconds"] = round(time.time() - t_start, 2)
    log_entry["scraped_at"] = datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

    return normalized, log_entry


# ── Generator (used by both CLI and FastAPI) ──────────────────────────────────

MAX_WORKERS = 6  # Parallel scrape threads (polite but fast)


def _scrape_roaster_task(idx, roaster, total):
    """
    Worker function for parallel scraping. Returns a result dict
    that the generator can yield as an event.
    """
    name = roaster["name"]
    platform = roaster.get("platform", "custom").lower()

    try:
        normalized, log_entry = _scrape_single_roaster(roaster)
        detected = log_entry.get("platform_detected", platform)
        return {
            "idx": idx,
            "success": True,
            "roaster": name,
            "platform": detected,
            "products": normalized,
            "log_entry": log_entry,
        }
    except Exception as exc:
        err = str(exc)[:200]
        log_entry = {
            "roaster": name,
            "website": roaster["website"],
            "platform_catalog": platform,
            "platform_detected": None,
            "status": "failed",
            "error": err,
            "scrape_duration_seconds": 0.0,
            "scraped_at": datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        }
        return {
            "idx": idx,
            "success": False,
            "roaster": name,
            "error": err,
            "log_entry": log_entry,
        }


def scrape_all_generator(catalog_path=None):
    """
    Generator that yields event dicts as each roaster is processed.
    Uses a thread pool (MAX_WORKERS threads) for parallel scraping.
    Results are yielded as they complete (not necessarily in order).
    """
    roasters = _load_catalog(catalog_path)
    all_products = []
    scrape_log = []
    total = len(roasters)
    done_count = 0

    # Submit all roasters to the thread pool
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(_scrape_roaster_task, idx, roaster, total): (idx, roaster)
            for idx, roaster in enumerate(roasters, start=1)
        }

        for future in as_completed(futures):
            result = future.result()
            done_count += 1
            name = result["roaster"]

            if result["success"]:
                normalized = result["products"]
                all_products.extend(normalized)
                scrape_log.append(result["log_entry"])

                yield {
                    "event": "roaster_done",
                    "data": {
                        "index": done_count,
                        "total": total,
                        "roaster": name,
                        "platform": result["platform"],
                        "coffees_found": len(normalized),
                        "products": normalized,
                    },
                }

                print(
                    f"[{done_count}/{total}] {name} ({result['platform']}) "
                    f"... {len(normalized)} coffees ✓"
                )

            else:
                scrape_log.append(result["log_entry"])

                yield {
                    "event": "roaster_failed",
                    "data": {
                        "index": done_count,
                        "total": total,
                        "roaster": name,
                        "error": result["error"],
                    },
                }

                print(f"[{done_count}/{total}] {name} ... FAILED: {result['error']} ✗")

    # Finalize
    all_products = _dedup_products(all_products)

    print("\nWriting output files...")
    _write_outputs(all_products, scrape_log)

    succeeded = sum(1 for e in scrape_log if e["status"] == "success")
    failed = sum(1 for e in scrape_log if e["status"] == "failed")

    summary = {
        "total_products": len(all_products),
        "total_roasters": total,
        "succeeded": succeeded,
        "failed": failed,
    }

    print(f"\n{'═' * 55}")
    print("SCRAPE COMPLETE")
    print(f"  Roasters attempted : {total}")
    print(f"  Roasters succeeded : {succeeded}")
    print(f"  Roasters failed    : {failed}")
    print(f"  Total products     : {len(all_products)}")
    print(f"{'═' * 55}")

    yield {"event": "scrape_complete", "data": summary}


# ── CLI entry point ───────────────────────────────────────────────────────────

def main():
    """Run the full pipeline from the command line."""
    os.makedirs(_OUTPUT_DIR, exist_ok=True)
    # Consume the generator to completion
    for _event in scrape_all_generator():
        pass


if __name__ == "__main__":
    main()
